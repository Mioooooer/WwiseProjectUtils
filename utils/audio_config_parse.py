import openpyxl
import openpyxl.cell

def parseMergedCell(ws, cell):
    if isinstance(cell, openpyxl.cell.MergedCell):
        for mergedRange in ws.merged_cells.ranges:
            if cell.coordinate in mergedRange:
                cell = ws.cell(row = mergedRange.min_row, column = mergedRange.min_col)
                break
    return cell

def concatenatePrefix(pA, pB, concatSymbol):
    pA = pA.strip(concatSymbol)
    pB = pB.strip(concatSymbol)
    if pA == "":
        return pB
    if pB == "":
        return pA
    return pA + concatSymbol + pB

def getValueFromDict(valueDict, key):
    while key != "":
        value = valueDict.get(key)
        if value == None:
            key = key.split("_")
            key.pop()
            key = "_".join(key)
        else:
            return value
    return None

def getRuleFromFile(xlsxpath, wsName):#wsName for example: "LoudnessRule"
    LoudnessRuleDict = {}
    wb = openpyxl.load_workbook(xlsxpath)
    RuleWs = wb[wsName]
    for rowIndex in range(1, RuleWs.max_row + 1):
        concatStr = ""
        colIndex = 1
        targetCell = RuleWs.cell(rowIndex, colIndex)
        cell = parseMergedCell(RuleWs, targetCell)
        valueFlag = False
        while cell.value != None:
            if not "," in cell.value:
                if not valueFlag:
                    if not concatStr.endswith(cell.value):
                        concatStr = concatenatePrefix(concatStr, cell.value, "_")
                else:
                    if concatStr != "":
                        concatStr = concatStr.split("_")
                        concatStr[-1] = cell.value
                        concatStr = "_".join(concatStr)
            else:
                valueFlag = True
                valueSplit = cell.value.split(",")
                if valueSplit[0] != "" and valueSplit[1] != "":
                    if concatStr != "":
                        valueMin = min(float(valueSplit[0]), float(valueSplit[1]))
                        valueMax = max(float(valueSplit[0]), float(valueSplit[1]))
                        if valueMax == valueMin:
                            valueMin = None
                        LoudnessRuleDict[concatStr] = {"valueMin": valueMin, "valueMax": valueMax}
                elif valueSplit[0] == "" and valueSplit[1] != "":
                    LoudnessRuleDict[concatStr] = {"valueMin": None, "valueMax": float(valueSplit[1])}
                elif valueSplit[0] != "" and valueSplit[1] == "":
                    LoudnessRuleDict[concatStr] = {"valueMin": float(valueSplit[0]), "valueMax": 0}
            colIndex += 1
            targetCell = RuleWs.cell(rowIndex, colIndex)
            cell = parseMergedCell(RuleWs, targetCell)
    return LoudnessRuleDict